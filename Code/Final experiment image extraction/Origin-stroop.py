import matplotlib.pyplot as plt
from tkinter import Tk
from tkinter.filedialog import askopenfilename, askdirectory
from openpyxl import load_workbook

# 定义背景颜色（归一化RGB），此处为灰色
bg_color = (200/255, 200/255, 200/255)

# 隐藏Tkinter主窗口
root = Tk()
root.withdraw()

# 弹出文件对话框选择Excel文件
file_path = askopenfilename(title="请选择Excel文件", filetypes=[("Excel files", "*.xlsx;*.xls")])
if not file_path:
    print("未选择Excel文件，程序退出。")
    exit()

# 弹出文件夹选择对话框，选择保存图片的文件夹
output_folder = askdirectory(title="请选择保存图片的文件夹")
if not output_folder:
    print("未选择保存路径，程序退出。")
    exit()

# 使用 openpyxl 加载Excel数据（保留格式信息）
wb = load_workbook(file_path)
ws = wb.active

# 构造表头映射（假设第一行是表头）
header_to_col = {}
for cell in ws[1]:
    header_to_col[cell.value] = cell.column

# 固定使用的列名称：Target 用于生成图片内容，Image 用于图片文件名
target_col_name = 'Target'
image_col_name = 'Image'

# 获取对应的列号
target_col_idx = header_to_col.get(target_col_name)
image_col_idx = header_to_col.get(image_col_name)

if None in [target_col_idx, image_col_idx]:
    print("错误：找不到指定的 Target 或 Image 列，请检查表头名称。")
    exit()

# 输入行号范围（注意：数据行编号从1开始，表头除外）
start_row = int(input("请输入起始行号（从1开始，表头除外）："))
end_row = int(input("请输入结束行号："))

def get_font_color(cell):
    """
    从 openpyxl 单元格中获取字体颜色，返回 hex 字符串供 matplotlib 使用。
    如果未设置字体颜色，则默认返回黑色 "#000000"。
    """
    if cell.font and cell.font.color and cell.font.color.type == 'rgb' and cell.font.color.rgb:
        rgb = cell.font.color.rgb  # 格式通常为 'FFRRGGBB'
        if len(rgb) == 8:
            return "#" + rgb[2:]
        else:
            return "#" + rgb
    return "#000000"  # 默认返回黑色

# 针对指定数据行生成图片
for i in range(start_row, end_row + 1):
    excel_row = i + 1  # Excel中实际行号（第一行为表头）
    
    # 获取对应单元格对象
    cell_target = ws.cell(row=excel_row, column=target_col_idx)
    cell_image = ws.cell(row=excel_row, column=image_col_idx)
    
    # 读取 Target 文本，若为空则置为空字符串，并去除空白
    text_target = str(cell_target.value).strip() if cell_target.value is not None else ""
    # 读取 Image 列作为文件名
    raw_filename = str(cell_image.value).strip() if cell_image.value is not None else ""
    
    # 对纯数字文本进行补零处理（假设应为5位）
    if text_target.isdigit():
        text_target = text_target.zfill(5)
    if raw_filename.isdigit():
        raw_filename = raw_filename.zfill(5)
    filename = raw_filename + ".png"
    
    # 获取 Target 单元格的字体颜色（默认黑色）
    color_target = get_font_color(cell_target)
    
    # 创建图像，尺寸500×300像素（figsize=(5,3)，dpi=100），使用指定背景色
    fig = plt.figure(figsize=(5, 3), dpi=100, facecolor=bg_color)
    
    # 调整子图边距，确保没有多余空白
    plt.subplots_adjust(left=0, right=1, top=1, bottom=0)
    
    ax = fig.add_subplot(111)
    ax.set_facecolor(bg_color)
    plt.axis('off')  # 不显示坐标轴

    # 绘制文本：居中显示 Target 文本，字号设为40（可根据需求调整）
    plt.text(0.5, 0.5, text_target, ha='center', va='center', fontsize=45, color=color_target)

    # 拼接图片保存的完整路径
    save_path = output_folder + "/" + filename

    # 保存图片（不使用 bbox_inches='tight' 保证输出尺寸为 500×300）
    plt.savefig(save_path, facecolor=fig.get_facecolor())
    plt.close()
    print(f"生成图片：{save_path}")
