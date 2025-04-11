import pandas as pd
import matplotlib.pyplot as plt
from tkinter import Tk
from tkinter.filedialog import askopenfilename, askdirectory
from openpyxl import load_workbook

# 定义背景颜色（归一化RGB），此处为灰色
bg_color = (200/255, 200/255, 200/255)

# 隐藏Tkinter主窗口
root = Tk()
root.withdraw()

# 弹出文件对话框选择Excel文件
file_path = askopenfilename(title="请选择Excel文件", filetypes=[("Excel files", "*.xlsx;*.xls")])
if not file_path:
    print("未选择Excel文件，程序退出。")
    exit()

# 弹出文件夹选择对话框，选择保存图片的文件夹
output_folder = askdirectory(title="请选择保存图片的文件夹")
if not output_folder:
    print("未选择保存路径，程序退出。")
    exit()

# 使用 openpyxl 加载Excel数据（保留格式信息）
wb = load_workbook(file_path)
ws = wb.active

# 构造表头映射（假设第一行是表头）
header_to_col = {}
for cell in ws[1]:
    header_to_col[cell.value] = cell.column  # cell.column 返回列号（数字）

# 固定使用的列名称（注意大小写需与Excel中一致）
target_col_name = 'Target'
option_left_col_name = 'Option_Left'
option_right_col_name = 'Option_Right'
image_col_name = 'Image'

# 获取各列的列号
target_col_idx = header_to_col.get(target_col_name)
option_left_col_idx = header_to_col.get(option_left_col_name)
option_right_col_idx = header_to_col.get(option_right_col_name)
image_col_idx = header_to_col.get(image_col_name)

if None in [target_col_idx, option_left_col_idx, option_right_col_idx, image_col_idx]:
    print("错误：找不到指定的某一列，请检查表头名称。")
    exit()

# 输入行号范围（注意：此处输入的是数据行号，第一行表头不算在内；
# 因此Excel中实际数据行的行号 = 用户输入行号 + 1）
start_row = int(input("请输入起始行号（从1开始，表头除外）："))
end_row = int(input("请输入结束行号："))

def get_font_color(cell):
    """
    从 openpyxl 单元格中获取字体颜色，返回 hex 字符串供 matplotlib 使用。
    如果未设置字体颜色，则默认返回黑色 "#000000"。
    """
    if cell.font and cell.font.color and cell.font.color.type == 'rgb' and cell.font.color.rgb:
        rgb = cell.font.color.rgb  # 通常为 'FFRRGGBB' 格式
        if len(rgb) == 8:
            # 忽略前两位Alpha
            return "#" + rgb[2:]
        else:
            return "#" + rgb
    return "#000000"  # 默认返回黑色

# 针对指定数据行生成图片
for i in range(start_row, end_row + 1):
    excel_row = i + 1  # Excel中实际行号（第一行为表头）
    
    # 获取对应单元格对象
    cell_target = ws.cell(row=excel_row, column=target_col_idx)
    cell_option_left = ws.cell(row=excel_row, column=option_left_col_idx)
    cell_option_right = ws.cell(row=excel_row, column=option_right_col_idx)
    cell_image = ws.cell(row=excel_row, column=image_col_idx)
    
    # 读取文本值，若为空则置为空字符串，并去除空白
    text_target = str(cell_target.value).strip() if cell_target.value is not None else ""
    text_option_left = str(cell_option_left.value).strip() if cell_option_left.value is not None else ""
    text_option_right = str(cell_option_right.value).strip() if cell_option_right.value is not None else ""
    raw_filename = str(cell_image.value).strip() if cell_image.value is not None else ""
    
    # 对全数字的值进行补零（假设应为5位）
    if text_target.isdigit():
        text_target = text_target.zfill(5)
    if raw_filename.isdigit():
        raw_filename = raw_filename.zfill(5)
    filename = raw_filename + ".png"
    
    # 获取单元格的字体颜色（默认黑色）
    color_target = get_font_color(cell_target)
    color_option_left = get_font_color(cell_option_left)
    color_option_right = get_font_color(cell_option_right)
    
    # 创建图像，尺寸500x300像素（figsize=(5,3) 与 dpi=100），背景色为 bg_color
    fig = plt.figure(figsize=(5, 3), dpi=100, facecolor=bg_color)
    ax = fig.add_subplot(111)
    ax.set_facecolor('gray')
    plt.axis('off')  # 不显示坐标轴

    # 调整子图边距，确保整个画布都被保存
    plt.subplots_adjust(left=0, right=1, top=1, bottom=0)
    
    # 绘制文本：
    # 中上位置显示 Target 文本（坐标：0.5, 0.70）
    plt.text(0.5, 0.70, text_target, ha='center', va='center', fontsize=36, color=color_target)
    # 左下位置显示 Option_Left 文本（坐标：0.25, 0.40）
    plt.text(0.25, 0.40, text_option_left, ha='center', va='center', fontsize=32, color=color_option_left)
    # 右下位置显示 Option_Right 文本（坐标：0.75, 0.40）
    plt.text(0.75, 0.40, text_option_right, ha='center', va='center', fontsize=32, color=color_option_right)

    # 拼接生成的图片文件完整保存路径
    save_path = output_folder + "/" + filename

    # 保存图片时不使用 bbox_inches 参数，确保图片尺寸严格为 500x300 像素
    plt.savefig(save_path, facecolor=fig.get_facecolor())
    plt.close()
    print(f"生成图片：{save_path}")
