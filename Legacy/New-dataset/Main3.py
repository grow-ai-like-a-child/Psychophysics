import string
import os
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# ================= Global 设置 ===================
# 对于 letter 版本使用的颜色顺序（红、蓝、绿、黄、橙、黑、紫）
letter_colors = ["Red", "Blue", "Green", "Yellow", "Orange", "Black", "Purple"]
color_map = {
    "Red": "FF0000",     # 红
    "Blue": "0000FF",    # 蓝
    "Green": "00FF00",   # 绿
    "Yellow": "FFFF00",  # 黄
    "Orange": "FFA500",  # 橙
    "Black": "000000",   # 黑
    "Purple": "800080"   # 紫
}

# ================= Flanker - Letter 版本 数据集生成 ===================
def generate_flanker_letter_type1():
    """
    flanker letter type1（交换第二列和第三列）：
      原始生成每一行为 [first*5, first*5, second*5]，
      交换后输出为 [first*5, second*5, first*5]。
      共生成 26*25 = 650 行数据。
      ※此处为区分效果，第一列采用 first*6
    """
    rows = [["Title", "Wrong_Option", "Right_Option"]]
    letters = list(string.ascii_uppercase)
    for first in letters:
        for second in letters:
            if first == second:
                continue
            rows.append([first * 6, second * 5, first * 5])
    return rows

def generate_flanker_letter_type2():
    """
    flanker letter type2（保持不变）：
      每一行构造为：
         left  = fixed*2 + other + fixed*2    (例如 "AABAA")
         right = other*2 + fixed + other*2     (例如 "BBABB")
      输出为：[left, left, right]。
    """
    rows = [["Title", "Wrong_Option", "Right_Option"]]
    letters = list(string.ascii_uppercase)
    for fixed in letters:
        for other in letters:
            if fixed == other:
                continue
            left = fixed * 2 + other + fixed * 2
            right = other * 2 + fixed + other * 2
            rows.append([left, left, right])
    return rows

def generate_flanker_letter_type3():
    """
    flanker letter type3：
      生成：
         col1 = fixed*5                (例如 "AAAAA")
         col2 = fixed*2 + other + fixed*2  (例如 "AABAA")
         col3 = other*2 + fixed + other*2    (例如 "BBABB")
      输出顺序为 [col1, col2, col3]。
    """
    rows = [["Title", "Wrong_Option", "Right_Option"]]
    letters = list(string.ascii_uppercase)
    for fixed in letters:
        for other in letters:
            if fixed == other:
                continue
            col1 = fixed * 6
            col2 = fixed * 2 + other + fixed * 2
            col3 = other * 2 + fixed + other * 2
            rows.append([col1, col2, col3])
    return rows

def generate_flanker_letter_type4():
    """
    flanker letter type4：
      原始生成：
         left  = fixed*2 + other + fixed*2    (例如 "AABAA")
         mid   = fixed*5                     (例如 "AAAAA")
         right = other*5                     (例如 "BBBBB")
      交换后输出为 [left, right, mid]。
    """
    rows = [["Title", "Wrong_Option", "Right_Option"]]
    letters = list(string.ascii_uppercase)
    for fixed in letters:
        for other in letters:
            if fixed == other:
                continue
            left  = fixed * 2 + other + fixed * 2
            mid   = fixed * 5
            right = other * 5
            rows.append([left, right, mid])
    return rows

# ================= Flanker - Number 版本 数据集生成 ===================
def generate_flanker_number_type1():
    """
    flanker number type1（交换第二列和第三列）：
      输出为 [str(first)*5, str(second)*5, str(first)*5]。
    """
    rows = [["Title", "Wrong_Option", "Right_Option"]]
    for first in range(1, 10):
        for second in range(1, 10):
            if first == second:
                continue
            rows.append([str(first) * 5, str(second) * 5, str(first) * 5])
    return rows

def generate_flanker_number_type2():
    """
    flanker number type2（保持不变）：
      每一行构造为：
         left  = fixed*2 + other + fixed*2    (例如 "11211")
         right = other*2 + fixed + other*2     (例如 "22322")
      输出为：[left, left, right]。
    """
    rows = [["Title", "Wrong_Option", "Right_Option"]]
    digits = "123456789"
    for fixed in digits:
        for other in digits:
            if fixed == other:
                continue
            left = fixed * 2 + other + fixed * 2
            right = other * 2 + fixed + other * 2
            rows.append([left, left, right])
    return rows

def generate_flanker_number_type3():
    """
    flanker number type3：
      生成：
         col1 = fixed*5                 (例如 "11111")
         col2 = fixed*2 + other + fixed*2   (例如 "11211")
         col3 = other*2 + fixed + other*2     (例如 "22122")
      输出顺序为 [col1, col2, col3]。
    """
    rows = [["Title", "Wrong_Option", "Right_Option"]]
    digits = "123456789"
    for fixed in digits:
        for other in digits:
            if fixed == other:
                continue
            col1 = fixed * 5
            col2 = fixed * 2 + other + fixed * 2
            col3 = other * 2 + fixed + other * 2
            rows.append([col1, col2, col3])
    return rows

def generate_flanker_number_type4():
    """
    flanker number type4：
      生成：
         col1 = x*2 + y + x*2    (例如 "11211")
         col2 = y*5              (例如 "22222")
         col3 = x*5              (例如 "11111")
      输出顺序为 [col1, col2, col3]。
    """
    rows = [["Title", "Wrong_Option", "Right_Option"]]
    digits = "123456789"
    for x in digits:
        for y in digits:
            if x == y:
                continue
            col1 = x * 2 + y + x * 2
            col2 = y * 5
            col3 = x * 5
            rows.append([col1, col2, col3])
    return rows

# ================= Stroop - Letter 版本 数据集生成 ===================
def generate_stroop_letter_type1():
    wb = Workbook()
    ws = wb.active
    ws.title = "Color Sequences"
    ws.append(["Title", "Wrong_Option", "Right_Option"])
    row_idx = 2
    for A in letter_colors:
        for B in letter_colors:
            if A == B:
                continue
            seq = [A, B, A]
            for col_idx, word in enumerate(seq, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=word)
                if col_idx in [1, 3]:
                    font_color = color_map[A]
                else:
                    font_color = color_map[B]
                cell.font = Font(color=font_color)
            row_idx += 1
    base_folder = os.path.join(os.getcwd(), "Data-set")
    os.makedirs(base_folder, exist_ok=True)
    subfolder = os.path.join(base_folder, "Stroop-Letter")
    os.makedirs(subfolder, exist_ok=True)
    filename = "stroop_letter_type1_dataset.xlsx"
    filepath = os.path.join(subfolder, filename)
    wb.save(filepath)
    print(f"Stroop Letter Type1 数据集已保存到 {filepath}")
    return filepath

def generate_stroop_letter_type2():
    wb = Workbook()
    ws = wb.active
    ws.title = "Color Sequences"
    ws.append(["Title", "Wrong_Option", "Right_Option"])
    row_idx = 2
    for A in letter_colors:
        for B in letter_colors:
            if A == B:
                continue
            seq = [A, B, A]
            for col_idx, word in enumerate(seq, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=word)
                if col_idx in [1, 3]:
                    font_color = color_map[B]
                else:
                    font_color = color_map[A]
                cell.font = Font(color=font_color)
            row_idx += 1
    base_folder = os.path.join(os.getcwd(), "Data-set")
    os.makedirs(base_folder, exist_ok=True)
    subfolder = os.path.join(base_folder, "Stroop-Letter")
    os.makedirs(subfolder, exist_ok=True)
    filename = "stroop_letter_type2_dataset.xlsx"
    filepath = os.path.join(subfolder, filename)
    wb.save(filepath)
    print(f"Stroop Letter Type2 数据集已保存到 {filepath}")
    return filepath

def generate_stroop_letter_type3():
    wb = Workbook()
    ws = wb.active
    ws.title = "Color Sequences"
    ws.append(["Title", "Wrong_Option", "Right_Option"])
    row_idx = 2
    for A in letter_colors:
        for B in letter_colors:
            if A == B:
                continue
            seq = [A, B, A]
            for col_idx, word in enumerate(seq, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=word)
                if col_idx in [1, 2]:
                    font_color = color_map[A]
                else:
                    font_color = color_map[B]
                cell.font = Font(color=font_color)
            row_idx += 1
    base_folder = os.path.join(os.getcwd(), "Data-set")
    os.makedirs(base_folder, exist_ok=True)
    subfolder = os.path.join(base_folder, "Stroop-Letter")
    os.makedirs(subfolder, exist_ok=True)
    filename = "stroop_letter_type3_dataset.xlsx"
    filepath = os.path.join(subfolder, filename)
    wb.save(filepath)
    print(f"Stroop Letter Type3 数据集已保存到 {filepath}")
    return filepath

def generate_stroop_letter_type4():
    wb = Workbook()
    ws = wb.active
    ws.title = "Color Sequences"
    ws.append(["Title", "Wrong_Option", "Right_Option"])
    row_idx = 2
    for A in letter_colors:
        for B in letter_colors:
            if A == B:
                continue
            seq = [A, B, A]
            for col_idx, word in enumerate(seq, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=word)
                if col_idx in [1, 2]:
                    font_color = color_map[B]
                else:
                    font_color = color_map[A]
                cell.font = Font(color=font_color)
            row_idx += 1
    base_folder = os.path.join(os.getcwd(), "Data-set")
    os.makedirs(base_folder, exist_ok=True)
    subfolder = os.path.join(base_folder, "Stroop-Letter")
    os.makedirs(subfolder, exist_ok=True)
    filename = "stroop_letter_type4_dataset.xlsx"
    filepath = os.path.join(subfolder, filename)
    wb.save(filepath)
    print(f"Stroop Letter Type4 数据集已保存到 {filepath}")
    return filepath

# ================= 保存 Excel (用于 Flanker 数据集，使用 pandas) ===================
def save_to_excel(rows, task, style, flanker_type):
    """
    将生成的数据保存到 Excel 文件中，
    文件保存路径：Data-set/Flanker-<Style>/，
    Excel 文件名格式：<task>_<style>_<flanker_type>_dataset.xlsx
    列标题：Title, Wrong_Option, Right_Option
    """
    base_folder = os.path.join(os.getcwd(), "Data-set")
    os.makedirs(base_folder, exist_ok=True)
    subfolder = os.path.join(base_folder, f"{task.capitalize()}-{style.capitalize()}")
    os.makedirs(subfolder, exist_ok=True)
    filename = f"{task}_{style}_{flanker_type}_dataset.xlsx"
    filepath = os.path.join(subfolder, filename)
    df = pd.DataFrame(rows, columns=["Title", "Wrong_Option", "Right_Option"])
    df.to_excel(filepath, index=False)
    print(f"数据已保存到 Excel 文件：{filepath}")
    return filepath

# ================= 生成 Flanker 图片 (参考提供的代码) ===================
def generate_flanker_images_from_excel(excel_file, task, style, flanker_type):
    """
    根据 Excel 文件生成 Flanker 图片，
    图片保存在：Data-set-Image/Flanker-<Style>-image/<flanker_type>/，
    图片生成：每行 Title 为文本1，
      Wrong_Option 为文本2（显示在左下），
      Right_Option 为文本3（显示在右下）。
    此版本直接使用 pandas 读取 Excel 文件，自动对除表头外所有数据行生成图片，
    图片背景色为灰色，尺寸500x300像素（figsize=(5,3)，dpi=100）。
    文件名以数字序号（从1开始）开头。
    """
    df = pd.read_excel(excel_file)
    base_folder = os.path.join(os.getcwd(), "Data-set-Image")
    os.makedirs(base_folder, exist_ok=True)
    subfolder = os.path.join(base_folder, f"{task.capitalize()}-{style.capitalize()}-image")
    os.makedirs(subfolder, exist_ok=True)
    output_folder = os.path.join(subfolder, flanker_type)
    os.makedirs(output_folder, exist_ok=True)
    
    print("图片生成中...")
    count = 1
    # 遍历除表头外所有数据行
    for idx, row in df.iloc[1:].iterrows():
        text_top = str(row["Title"]).strip() if pd.notnull(row["Title"]) else ""
        text_wrong = str(row["Wrong_Option"]).strip() if pd.notnull(row["Wrong_Option"]) else ""
        text_right = str(row["Right_Option"]).strip() if pd.notnull(row["Right_Option"]) else ""
        
        fig = plt.figure(figsize=(5, 3), dpi=100, facecolor='gray')
        ax = fig.add_subplot(111)
        ax.set_facecolor('gray')
        plt.axis('off')
        
        # 文本1：Title 放在中上 (x=0.5, y=0.70)
        plt.text(0.5, 0.70, text_top, ha='center', va='center', fontsize=25, color='white')
        # 文本2：Wrong_Option 放在左下 (x=0.25, y=0.40)
        plt.text(0.25, 0.40, text_wrong, ha='center', va='center', fontsize=25, color='white')
        # 文本3：Right_Option 放在右下 (x=0.75, y=0.40)
        plt.text(0.75, 0.40, text_right, ha='center', va='center', fontsize=25, color='white')
        
        filename = f"{count}_{text_top}_{text_wrong}_{text_right}.png"
        for ch in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
            filename = filename.replace(ch, '_')
        save_path = os.path.join(output_folder, filename)
        plt.savefig(save_path, bbox_inches='tight', pad_inches=0.1, facecolor='gray')
        plt.close()
        count += 1
    print(f"图片生成中... 共计 {count - 1} 张。")

# ================= 生成 Stroop 图片 (保持不变) ===================
def get_font_color(cell):
    """
    读取单元格字体颜色，返回标准的十六进制颜色代码（如 "#RRGGBB"）
    """
    color_hex = "#000000"  # 默认黑色
    if cell.font and cell.font.color and cell.font.color.rgb:
        try:
            rgb_val = cell.font.color.rgb
            if not isinstance(rgb_val, str):
                rgb_str = str(rgb_val)
            else:
                rgb_str = rgb_val
            if len(rgb_str) == 8 and all(ch in "0123456789ABCDEFabcdef" for ch in rgb_str):
                color_hex = "#" + rgb_str[2:]
            elif len(rgb_str) == 6 and all(ch in "0123456789ABCDEFabcdef" for ch in rgb_str):
                color_hex = "#" + rgb_str
        except Exception as e:
            color_hex = "#000000"
    return color_hex

def generate_stroop_images_from_excel(excel_file, task, style, dataset_type):
    """
    根据 Excel 文件生成 Stroop 图片，
    图片保存在：Data-set-Image/Stroop-Letter-image/<dataset_type>/，
    使用 openpyxl 读取 Excel（保留字体颜色信息），
    文本位置：
      Title 放在中上 (0.5,0.70, fontsize=30)；
      Wrong_Option 放在右下 (0.75,0.40, fontsize=25)；
      Right_Option 放在左下 (0.25,0.40, fontsize=25)。
    背景色为归一化灰色：(200/255,200/255,200/255)，
    文件名以数字序号（从1开始）开头，自动生成所有数据行的图片。
    """
    wb = load_workbook(excel_file)
    ws = wb.active
    base_folder = os.path.join(os.getcwd(), "Data-set-Image")
    os.makedirs(base_folder, exist_ok=True)
    subfolder = os.path.join(base_folder, f"{task.capitalize()}-{style.capitalize()}-image")
    os.makedirs(subfolder, exist_ok=True)
    output_folder = os.path.join(subfolder, dataset_type)
    os.makedirs(output_folder, exist_ok=True)
    
    bg_color = (200/255, 200/255, 200/255)
    print("图片生成中...")
    count = 1
    for r in range(2, ws.max_row + 1):
        cell_top = ws[f"A{r}"]
        cell_wrong = ws[f"B{r}"]
        cell_right = ws[f"C{r}"]
        text_top = str(cell_top.value).strip() if cell_top.value is not None else ""
        text_wrong = str(cell_wrong.value).strip() if cell_wrong.value is not None else ""
        text_right = str(cell_right.value).strip() if cell_right.value is not None else ""
        
        color_top = get_font_color(cell_top)
        color_wrong = get_font_color(cell_wrong)
        color_right = get_font_color(cell_right)
        
        fig = plt.figure(figsize=(5, 3), dpi=100, facecolor=bg_color)
        plt.axis('off')
        
        plt.text(0.5, 0.70, text_top, ha='center', va='center', fontsize=30, color=color_top)
        plt.text(0.75, 0.40, text_wrong, ha='center', va='center', fontsize=25, color=color_wrong)
        plt.text(0.25, 0.40, text_right, ha='center', va='center', fontsize=25, color=color_right)
        
        filename = f"{count}_{text_top}_{text_wrong}_{text_right}.png"
        for ch in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
            filename = filename.replace(ch, '_')
        save_path = os.path.join(output_folder, filename)
        plt.savefig(save_path, bbox_inches='tight', pad_inches=0.1, facecolor=bg_color)
        plt.close()
        count += 1
    print(f"图片生成中... 共计 {count - 1} 张。")

# ================= 主程序 ===================
def main():
    # 选择任务：flanker 或 stroop
    task = input("请选择任务（输入 flanker 或 stroop）：").strip().lower()
    if task not in ["flanker", "stroop"]:
        print("未知任务。")
        return
    if task == "stroop":
        style = "letter"  # stroop 默认使用 letter 版本
    else:
        style = input("请选择数据类型（输入 letter 或 number）：").strip().lower()
        if style not in ["letter", "number"]:
            print("请选择 letter 或 number。")
            return
    dataset_type = input("请选择类型（例如 type1/type2/type3/type4）：").strip().lower()
    rows = None
    excel_filepath = ""
    
    if task == "flanker":
        if style == "letter":
            if dataset_type == "type1":
                rows = generate_flanker_letter_type1()
                print("生成 flanker letter type1 数据集。")
            elif dataset_type == "type2":
                rows = generate_flanker_letter_type2()
                print("生成 flanker letter type2 数据集。")
            elif dataset_type == "type3":
                rows = generate_flanker_letter_type3()
                print("生成 flanker letter type3 数据集。")
            elif dataset_type == "type4":
                rows = generate_flanker_letter_type4()
                print("生成 flanker letter type4 数据集。")
        elif style == "number":
            if dataset_type == "type1":
                rows = generate_flanker_number_type1()
                print("生成 flanker number type1 数据集。")
            elif dataset_type == "type2":
                rows = generate_flanker_number_type2()
                print("生成 flanker number type2 数据集。")
            elif dataset_type == "type3":
                rows = generate_flanker_number_type3()
                print("生成 flanker number type3 数据集。")
            elif dataset_type == "type4":
                rows = generate_flanker_number_type4()
                print("生成 flanker number type4 数据集。")
        excel_filepath = save_to_excel(rows, task, style, dataset_type)
    elif task == "stroop":
        if dataset_type == "type1":
            excel_filepath = generate_stroop_letter_type1()
        elif dataset_type == "type2":
            excel_filepath = generate_stroop_letter_type2()
        elif dataset_type == "type3":
            excel_filepath = generate_stroop_letter_type3()
        elif dataset_type == "type4":
            excel_filepath = generate_stroop_letter_type4()
        else:
            print("未知 stroop 类型。")
            return

    gen_img = input("是否生成图像？(y/n)：").strip().lower()
    if gen_img == "y":
        if task == "flanker":
            generate_flanker_images_from_excel(excel_filepath, task, style, dataset_type)
        else:
            generate_stroop_images_from_excel(excel_filepath, task, style, dataset_type)
    else:
        print("不生成图像。")

if __name__ == "__main__":
    main()
