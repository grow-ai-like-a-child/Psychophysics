import string
import os
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# ================= Global Settings ===================
# For the letter version, use the following color order (Red, Blue, Green, Yellow, Orange, Black, Purple)
letter_colors = ["Red", "Blue", "Green", "Yellow", "Orange", "Black", "Purple"]
color_map = {
    "Red": "FF0000",     # Red
    "Blue": "0000FF",    # Blue
    "Green": "00FF00",   # Green
    "Yellow": "FFFF00",  # Yellow
    "Orange": "FFA500",  # Orange
    "Black": "000000",   # Black
    "Purple": "800080"   # Purple
}

# ================= Flanker - Letter Version Data Generation ===================
def generate_flanker_letter_type1():
    """
    Flanker letter type1 (swap second and third column):
      Originally, each row is generated as [first*5, first*5, second*5],
      then swapped to output as [first*5, second*5, first*5].
      Total generated: 26*25 = 650 rows.
      Note: Here, for demonstration, the first column is first*6.
    """
    rows = [["Title", "Wrong_Option", "Right_Option"]]
    letters = list(string.ascii_uppercase)
    for first in letters:
        for second in letters:
            if first == second:
                continue
            rows.append([first * 6, second * 5, first * 5])
    return rows

def generate_flanker_letter_type2():
    """
    Flanker letter type2 (unchanged):
      Each row is constructed as:
         left  = fixed*2 + other + fixed*2 (e.g., "AABAA")
         right = other*2 + fixed + other*2 (e.g., "BBABB")
      Output as: [left, left, right].
    """
    rows = [["Title", "Wrong_Option", "Right_Option"]]
    letters = list(string.ascii_uppercase)
    for fixed in letters:
        for other in letters:
            if fixed == other:
                continue
            left = fixed * 2 + other + fixed * 2
            right = other * 2 + fixed + other * 2
            rows.append([left, left, right])
    return rows

def generate_flanker_letter_type3():
    """
    Flanker letter type3:
      Generate:
         col1 = fixed*5 (e.g., "AAAAA")
         col2 = fixed*2 + other + fixed*2 (e.g., "AABAA")
         col3 = other*2 + fixed + other*2 (e.g., "BBABB")
      Output order: [col1, col2, col3],
      i.e., generates "AAAAA    AABAA    BBABB".
    """
    rows = [["Title", "Wrong_Option", "Right_Option"]]
    letters = list(string.ascii_uppercase)
    for fixed in letters:
        for other in letters:
            if fixed == other:
                continue
            col1 = fixed * 6
            col2 = fixed * 2 + other + fixed * 2
            col3 = other * 2 + fixed + other * 2
            rows.append([col1, col2, col3])
    return rows

def generate_flanker_letter_type4():
    """
    Flanker letter type4:
      Originally generate:
         left  = fixed*2 + other + fixed*2 (e.g., "AABAA")
         mid   = fixed*5                  (e.g., "AAAAA")
         right = other*5                  (e.g., "BBBBB")
      Then swap to output as [left, right, mid].
    """
    rows = [["Title", "Wrong_Option", "Right_Option"]]
    letters = list(string.ascii_uppercase)
    for fixed in letters:
        for other in letters:
            if fixed == other:
                continue
            left  = fixed * 2 + other + fixed * 2
            mid   = fixed * 5
            right = other * 5
            rows.append([left, right, mid])
    return rows

# ================= Flanker - Number Version Data Generation ===================
def generate_flanker_number_type1():
    """
    Flanker number type1 (swap second and third column):
      Output as [str(first)*5, str(second)*5, str(first)*5].
    """
    rows = [["Title", "Wrong_Option", "Right_Option"]]
    for first in range(1, 10):
        for second in range(1, 10):
            if first == second:
                continue
            rows.append([str(first) * 5, str(second) * 5, str(first) * 5])
    return rows

def generate_flanker_number_type2():
    """
    Flanker number type2 (unchanged):
      Each row is constructed as:
         left  = fixed*2 + other + fixed*2 (e.g., "11211")
         right = other*2 + fixed + other*2  (e.g., "22322")
      Output as: [left, left, right].
    """
    rows = [["Title", "Wrong_Option", "Right_Option"]]
    digits = "123456789"
    for fixed in digits:
        for other in digits:
            if fixed == other:
                continue
            left = fixed * 2 + other + fixed * 2
            right = other * 2 + fixed + other * 2
            rows.append([left, left, right])
    return rows

def generate_flanker_number_type3():
    """
    Flanker number type3:
      Generate:
         col1 = fixed*5 (e.g., "11111")
         col2 = fixed*2 + other + fixed*2 (e.g., "11211")
         col3 = other*2 + fixed + other*2 (e.g., "22122")
      Output order: [col1, col2, col3].
    """
    rows = [["Title", "Wrong_Option", "Right_Option"]]
    digits = "123456789"
    for fixed in digits:
        for other in digits:
            if fixed == other:
                continue
            col1 = fixed * 5
            col2 = fixed * 2 + other + fixed * 2
            col3 = other * 2 + fixed + other * 2
            rows.append([col1, col2, col3])
    return rows

def generate_flanker_number_type4():
    """
    Flanker number type4:
      Generate:
         col1 = x*2 + y + x*2 (e.g., "11211")
         col2 = y*5           (e.g., "22222")
         col3 = x*5           (e.g., "11111")
      Output order: [col1, col2, col3].
    """
    rows = [["Title", "Wrong_Option", "Right_Option"]]
    digits = "123456789"
    for x in digits:
        for y in digits:
            if x == y:
                continue
            col1 = x * 2 + y + x * 2
            col2 = y * 5
            col3 = x * 5
            rows.append([col1, col2, col3])
    return rows

# ================= Stroop - Letter Version Data Generation ===================
def generate_stroop_letter_type1():
    wb = Workbook()
    ws = wb.active
    ws.title = "Color Sequences"
    ws.append(["Title", "Wrong_Option", "Right_Option"])
    row_idx = 2
    for A in letter_colors:
        for B in letter_colors:
            if A == B:
                continue
            seq = [A, B, A]
            for col_idx, word in enumerate(seq, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=word)
                # Stroop type1: Use color_map[A] for columns 1 and 3, and color_map[B] for column 2.
                if col_idx in [1, 3]:
                    font_color = color_map[A]
                else:
                    font_color = color_map[B]
                cell.font = Font(color=font_color)
            row_idx += 1
    base_folder = os.path.join(os.getcwd(), "Data-set")
    os.makedirs(base_folder, exist_ok=True)
    subfolder = os.path.join(base_folder, "Stroop-Letter")
    os.makedirs(subfolder, exist_ok=True)
    filename = "stroop_letter_type1_dataset.xlsx"
    filepath = os.path.join(subfolder, filename)
    ws.parent.save(filepath)
    print(f"Stroop Letter Type1 dataset saved to {filepath}")
    return filepath

def generate_stroop_letter_type2():
    wb = Workbook()
    ws = wb.active
    ws.title = "Color Sequences"
    ws.append(["Title", "Wrong_Option", "Right_Option"])
    row_idx = 2
    for A in letter_colors:
        for B in letter_colors:
            if A == B:
                continue
            seq = [A, B, A]
            for col_idx, word in enumerate(seq, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=word)
                # Stroop type2: Use color_map[B] for columns 1 and 3, and color_map[A] for column 2.
                if col_idx in [1, 3]:
                    font_color = color_map[B]
                else:
                    font_color = color_map[A]
                cell.font = Font(color=font_color)
            row_idx += 1
    base_folder = os.path.join(os.getcwd(), "Data-set")
    os.makedirs(base_folder, exist_ok=True)
    subfolder = os.path.join(base_folder, "Stroop-Letter")
    os.makedirs(subfolder, exist_ok=True)
    filename = "stroop_letter_type2_dataset.xlsx"
    filepath = os.path.join(subfolder, filename)
    ws.parent.save(filepath)
    print(f"Stroop Letter Type2 dataset saved to {filepath}")
    return filepath

def generate_stroop_letter_type3():
    wb = Workbook()
    ws = wb.active
    ws.title = "Color Sequences"
    ws.append(["Title", "Wrong_Option", "Right_Option"])
    row_idx = 2
    for A in letter_colors:
        for B in letter_colors:
            if A == B:
                continue
            seq = [A, B, A]
            for col_idx, word in enumerate(seq, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=word)
                # Stroop type3: Use color_map[A] for columns 1 and 2, and color_map[B] for column 3.
                if col_idx in [1, 2]:
                    font_color = color_map[A]
                else:
                    font_color = color_map[B]
                cell.font = Font(color=font_color)
            row_idx += 1
    base_folder = os.path.join(os.getcwd(), "Data-set")
    os.makedirs(base_folder, exist_ok=True)
    subfolder = os.path.join(base_folder, "Stroop-Letter")
    os.makedirs(subfolder, exist_ok=True)
    filename = "stroop_letter_type3_dataset.xlsx"
    filepath = os.path.join(subfolder, filename)
    ws.parent.save(filepath)
    print(f"Stroop Letter Type3 dataset saved to {filepath}")
    return filepath

def generate_stroop_letter_type4():
    wb = Workbook()
    ws = wb.active
    ws.title = "Color Sequences"
    ws.append(["Title", "Wrong_Option", "Right_Option"])
    row_idx = 2
    for A in letter_colors:
        for B in letter_colors:
            if A == B:
                continue
            seq = [A, B, A]
            for col_idx, word in enumerate(seq, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=word)
                # Stroop type4: Use color_map[B] for columns 1 and 2, and color_map[A] for column 3.
                if col_idx in [1, 2]:
                    font_color = color_map[B]
                else:
                    font_color = color_map[A]
                cell.font = Font(color=font_color)
            row_idx += 1
    base_folder = os.path.join(os.getcwd(), "Data-set")
    os.makedirs(base_folder, exist_ok=True)
    subfolder = os.path.join(base_folder, "Stroop-Letter")
    os.makedirs(subfolder, exist_ok=True)
    filename = "stroop_letter_type4_dataset.xlsx"
    filepath = os.path.join(subfolder, filename)
    ws.parent.save(filepath)
    print(f"Stroop Letter Type4 dataset saved to {filepath}")
    return filepath

# ================= Save Excel (for Flanker dataset using pandas) ===================
def save_to_excel(rows, task, style, flanker_type):
    """
    Save the generated data to an Excel file.
    File path: Data-set/Flanker-<Style>/,
    File name format: <task>_<style>_<flanker_type>_dataset.xlsx
    Column headers: Title, Wrong_Option, Right_Option
    """
    base_folder = os.path.join(os.getcwd(), "Data-set")
    os.makedirs(base_folder, exist_ok=True)
    subfolder = os.path.join(base_folder, f"{task.capitalize()}-{style.capitalize()}")
    os.makedirs(subfolder, exist_ok=True)
    filename = f"{task}_{style}_{flanker_type}_dataset.xlsx"
    filepath = os.path.join(subfolder, filename)
    df = pd.DataFrame(rows, columns=["Title", "Wrong_Option", "Right_Option"])
    df.to_excel(filepath, index=False)
    print(f"Data saved to Excel file: {filepath}")
    return filepath

# ================= Generate Flanker Images (based on provided code) ===================
def generate_flanker_images_from_excel(excel_file, task, style, flanker_type):
    """
    Generate Flanker images based on the Excel file.
    Save images to: Data-set-Image/Flanker-<Style>-image/<flanker_type>/
    For each row, display:
      Title as text1,
      Wrong_Option as text2 (displayed at bottom left),
      Right_Option as text3 (displayed at bottom right).
    This version uses pandas to read the Excel file and automatically generates images for all data rows (excluding header).
    Background color is gray, image size is 500x300 pixels (figsize=(5,3), dpi=100).
    File names start with a number (starting from 1).
    """
    df = pd.read_excel(excel_file)
    base_folder = os.path.join(os.getcwd(), "Data-set-Image")
    os.makedirs(base_folder, exist_ok=True)
    subfolder = os.path.join(base_folder, f"{task.capitalize()}-{style.capitalize()}-image")
    os.makedirs(subfolder, exist_ok=True)
    output_folder = os.path.join(subfolder, flanker_type)
    os.makedirs(output_folder, exist_ok=True)
    
    print("Generating images...")
    count = 1
    # Iterate over all rows except header
    for idx, row in df.iloc[1:].iterrows():
        text_top = str(row["Title"]).strip() if pd.notnull(row["Title"]) else ""
        text_wrong = str(row["Wrong_Option"]).strip() if pd.notnull(row["Wrong_Option"]) else ""
        text_right = str(row["Right_Option"]).strip() if pd.notnull(row["Right_Option"]) else ""
        
        fig = plt.figure(figsize=(5, 3), dpi=100, facecolor='gray')
        ax = fig.add_subplot(111)
        ax.set_facecolor('gray')
        plt.axis('off')
        
        # Text1: Title at center top (x=0.5, y=0.70)
        plt.text(0.5, 0.70, text_top, ha='center', va='center', fontsize=25, color='white')
        # Text2: Wrong_Option at bottom left (x=0.25, y=0.40)
        plt.text(0.25, 0.40, text_wrong, ha='center', va='center', fontsize=25, color='white')
        # Text3: Right_Option at bottom right (x=0.75, y=0.40)
        plt.text(0.75, 0.40, text_right, ha='center', va='center', fontsize=25, color='white')
        
        filename = f"{count}_{text_top}_{text_wrong}_{text_right}.png"
        for ch in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
            filename = filename.replace(ch, '_')
        save_path = os.path.join(output_folder, filename)
        plt.savefig(save_path, bbox_inches='tight', pad_inches=0.1, facecolor='gray')
        plt.close()
        count += 1
    print(f"Images generated... Total: {count - 1}")

# ================= Generate Stroop Images (unchanged) ===================
def get_font_color(cell):
    """
    Read cell font color and return standard hex color code (e.g., "#RRGGBB")
    """
    color_hex = "#000000"  # Default black
    if cell.font and cell.font.color and cell.font.color.rgb:
        try:
            rgb_val = cell.font.color.rgb
            if not isinstance(rgb_val, str):
                rgb_str = str(rgb_val)
            else:
                rgb_str = rgb_val
            if len(rgb_str) == 8 and all(ch in "0123456789ABCDEFabcdef" for ch in rgb_str):
                color_hex = "#" + rgb_str[2:]
            elif len(rgb_str) == 6 and all(ch in "0123456789ABCDEFabcdef" for ch in rgb_str):
                color_hex = "#" + rgb_str
        except Exception as e:
            color_hex = "#000000"
    return color_hex

def generate_stroop_images_from_excel(excel_file, task, style, dataset_type):
    """
    Generate Stroop images based on the Excel file.
    Save images to: Data-set-Image/Stroop-Letter-image/<dataset_type>/
    Uses openpyxl to read the Excel file (retaining font color information).
    Text positions:
      Title at center top (0.5, 0.70, fontsize=30);
      Wrong_Option at bottom right (0.75, 0.40, fontsize=25);
      Right_Option at bottom left (0.25, 0.40, fontsize=25).
    Background color is normalized gray: (200/255,200/255,200/255).
    File names start with a number (starting from 1).
    """
    wb = load_workbook(excel_file)
    ws = wb.active
    base_folder = os.path.join(os.getcwd(), "Data-set-Image")
    os.makedirs(base_folder, exist_ok=True)
    subfolder = os.path.join(base_folder, f"{task.capitalize()}-{style.capitalize()}-image")
    os.makedirs(subfolder, exist_ok=True)
    output_folder = os.path.join(subfolder, dataset_type)
    os.makedirs(output_folder, exist_ok=True)
    
    bg_color = (200/255, 200/255, 200/255)
    print("Generating images...")
    count = 1
    for r in range(2, ws.max_row + 1):
        cell_top = ws[f"A{r}"]
        cell_wrong = ws[f"B{r}"]
        cell_right = ws[f"C{r}"]
        text_top = str(cell_top.value).strip() if cell_top.value is not None else ""
        text_wrong = str(cell_wrong.value).strip() if cell_wrong.value is not None else ""
        text_right = str(cell_right.value).strip() if cell_right.value is not None else ""
        
        color_top = get_font_color(cell_top)
        color_wrong = get_font_color(cell_wrong)
        color_right = get_font_color(cell_right)
        
        fig = plt.figure(figsize=(5, 3), dpi=100, facecolor=bg_color)
        plt.axis('off')
        
        plt.text(0.5, 0.70, text_top, ha='center', va='center', fontsize=30, color=color_top)
        plt.text(0.75, 0.40, text_wrong, ha='center', va='center', fontsize=25, color=color_wrong)
        plt.text(0.25, 0.40, text_right, ha='center', va='center', fontsize=25, color=color_right)
        
        filename = f"{count}_{text_top}_{text_wrong}_{text_right}.png"
        for ch in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
            filename = filename.replace(ch, '_')
        save_path = os.path.join(output_folder, filename)
        plt.savefig(save_path, bbox_inches='tight', pad_inches=0.1, facecolor=bg_color)
        plt.close()
        count += 1
    print(f"Images generated... Total: {count - 1}")

# ================= Main Program ===================
def main():
    # Choose task: flanker or stroop
    task = input("Please choose task (enter flanker or stroop): ").strip().lower()
    if task not in ["flanker", "stroop"]:
        print("Unknown task.")
        return
    if task == "stroop":
        style = "letter"  # Stroop always uses letter version
    else:
        style = input("Please choose data type (enter letter or number): ").strip().lower()
        if style not in ["letter", "number"]:
            print("Please choose letter or number.")
            return
    dataset_type = input("Please choose type (e.g., type1/type2/type3/type4): ").strip().lower()
    rows = None
    excel_filepath = ""
    
    if task == "flanker":
        if style == "letter":
            if dataset_type == "type1":
                rows = generate_flanker_letter_type1()
                print("Flanker letter type1 dataset generated.")
            elif dataset_type == "type2":
                rows = generate_flanker_letter_type2()
                print("Flanker letter type2 dataset generated.")
            elif dataset_type == "type3":
                rows = generate_flanker_letter_type3()
                print("Flanker letter type3 dataset generated.")
            elif dataset_type == "type4":
                rows = generate_flanker_letter_type4()
                print("Flanker letter type4 dataset generated.")
        elif style == "number":
            if dataset_type == "type1":
                rows = generate_flanker_number_type1()
                print("Flanker number type1 dataset generated.")
            elif dataset_type == "type2":
                rows = generate_flanker_number_type2()
                print("Flanker number type2 dataset generated.")
            elif dataset_type == "type3":
                rows = generate_flanker_number_type3()
                print("Flanker number type3 dataset generated.")
            elif dataset_type == "type4":
                rows = generate_flanker_number_type4()
                print("Flanker number type4 dataset generated.")
        excel_filepath = save_to_excel(rows, task, style, dataset_type)
    elif task == "stroop":
        if dataset_type == "type1":
            excel_filepath = generate_stroop_letter_type1()
        elif dataset_type == "type2":
            excel_filepath = generate_stroop_letter_type2()
        elif dataset_type == "type3":
            excel_filepath = generate_stroop_letter_type3()
        elif dataset_type == "type4":
            excel_filepath = generate_stroop_letter_type4()
        else:
            print("Unknown stroop type.")
            return

    gen_img = input("Generate images? (y/n): ").strip().lower()
    if gen_img == "y":
        if task == "flanker":
            generate_flanker_images_from_excel(excel_filepath, task, style, dataset_type)
        else:
            generate_stroop_images_from_excel(excel_filepath, task, style, dataset_type)
    else:
        print("No images generated.")

if __name__ == "__main__":
    main()
